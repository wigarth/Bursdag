import datetime
from pathlib import Path

from openpyxl import load_workbook

# Filstier relativt til denne fila
BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "Bursdager.xlsx"   # merk stor B
HTML_PATH = BASE_DIR / "index.html"

BACKGROUND_BDAY = "Bilde.png"   # brukes når det ER bursdag
BACKGROUND_FALLBACK = "Bilde2.png"  # brukes når det IKKE er bursdag


def main() -> None:
    today = datetime.datetime.now()
    today_dm = today.strftime("%d.%m")  # f.eks. 06.02

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active  # bruker første ark

    bursdagsbarn = []

    # Forventer struktur:
    # Kolonne A: Fornavn
    # Kolonne B: Etternavn
    # Kolonne C: Fødselsdato (ekte dato)
    for row in ws.iter_rows(min_row=2, values_only=True):
        fornavn, etternavn, fodselsdato = row[:3]

        if not (fornavn and etternavn and fodselsdato):
            continue

        # fodselsdato bør være datetime.date eller datetime.datetime
        if not hasattr(fodselsdato, "strftime"):
            continue

        dm = fodselsdato.strftime("%d.%m")
        if dm == today_dm:
            bursdagsbarn.append(f"{fornavn} {etternavn}")

    if bursdagsbarn:
        # >>> DET ER BURSDAg(er)
        overskrift = "🎉 Gratulerer med dagen! 🎉"
        navn_html = "<br>".join(bursdagsbarn)

        html = f"""<!DOCTYPE html>
<html lang="no">
<head>
  <meta charset="UTF-8" />
  <title>BaRe Oslo – Bursdag</title>
  <style>
    html, body {{
      margin: 0;
      padding: 0;
      height: 100%;
      width: 100%;
    }}
    body {{
      font-family: Arial, sans-serif;
      color: white;
      text-align: center;
      background-image: url('{BACKGROUND_BDAY}');
      background-size: cover;
      background-position: center;
      background-repeat: no-repeat;
    }}
    .overlay {{
      background: rgba(0, 0, 0, 0.45);
      height: 100%;
      width: 100%;
      display: flex;
      flex-direction: column;
      justify-content: center;
      align-items: center;
      box-sizing: border-box;
      padding: 2rem;
    }}
    h1 {{
      font-size: 4rem;
      margin: 0 0 1rem 0;
    }}
    .names {{
      font-size: 3.5rem;
      font-weight: bold;
    }}
  </style>
</head>
<body>
  <div class="overlay">
    <h1>{overskrift}</h1>
    <div class="names">{navn_html}</div>
  </div>
</body>
</html>
"""
    else:
        # >>> INGEN bursdag – vis KUN Bilde2.png fullskjerm
        html = f"""<!DOCTYPE html>
<html lang="no">
<head>
  <meta charset="UTF-8" />
  <title>BaRe Oslo – Ingen bursdag</title>
  <style>
    html, body {{
      margin: 0;
      padding: 0;
      height: 100%;
      width: 100%;
    }}
    img {{
      width: 100%;
      height: 100%;
      object-fit: cover;
      display: block;
    }}
  </style>
</head>
<body>
  <img src="{BACKGROUND_FALLBACK}" alt="">
</body>
</html>
"""

    HTML_PATH.write_text(html, encoding="utf-8")


if __name__ == "__main__":
    main()
